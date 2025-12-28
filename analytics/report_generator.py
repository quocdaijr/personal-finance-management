"""
Report Generator Module - PDF and Excel report generation
"""
from datetime import datetime, timedelta
from typing import Dict, List, Any, Optional
import pandas as pd
from io import BytesIO
import os

# PDF generation
try:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import letter, A4
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from reportlab.lib.enums import TA_CENTER, TA_RIGHT
    REPORTLAB_AVAILABLE = True
except ImportError:
    REPORTLAB_AVAILABLE = False

# Excel generation
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.chart import BarChart, Reference, PieChart
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

from sqlalchemy import text
from sqlalchemy.orm import Session
from database import load_data_to_dataframe


class ReportGenerator:
    """Generate PDF and Excel reports"""

    def __init__(self, db: Session):
        self.db = db

    def generate_pdf_report(
        self,
        user_id: int,
        report_type: str,
        start_date: datetime,
        end_date: datetime,
        output_path: str
    ) -> Dict[str, Any]:
        """
        Generate PDF report

        Args:
            user_id: User ID
            report_type: Type of report (monthly, yearly, custom)
            start_date: Report start date
            end_date: Report end date
            output_path: Path to save PDF

        Returns:
            Dict containing generation result
        """
        if not REPORTLAB_AVAILABLE:
            return {
                "success": False,
                "error": "ReportLab not installed. Install with: pip install reportlab"
            }

        # Validate inputs
        if user_id <= 0:
            return {
                "success": False,
                "error": "Invalid user ID: must be positive"
            }

        if start_date > end_date:
            return {
                "success": False,
                "error": "Start date must be before end date"
            }

        # Validate output directory exists
        output_dir = os.path.dirname(output_path)
        if output_dir and not os.path.exists(output_dir):
            return {
                "success": False,
                "error": f"Output directory does not exist: {output_dir}"
            }

        try:
            # Create PDF document
            doc = SimpleDocTemplate(output_path, pagesize=letter)
            story = []
            styles = getSampleStyleSheet()

            # Title
            title_style = ParagraphStyle(
                'CustomTitle',
                parent=styles['Heading1'],
                fontSize=24,
                textColor=colors.HexColor('#1976d2'),
                spaceAfter=30,
                alignment=TA_CENTER
            )

            title = Paragraph(f"Financial Report - {report_type.replace('_', ' ').title()}", title_style)
            story.append(title)

            # Report period
            period_text = f"Period: {start_date.strftime('%Y-%m-%d')} to {end_date.strftime('%Y-%m-%d')}"
            story.append(Paragraph(period_text, styles['Normal']))
            story.append(Spacer(1, 0.3 * inch))

            # Get financial data
            data = self._get_report_data(user_id, start_date, end_date)

            # Summary section
            story.append(Paragraph("Financial Summary", styles['Heading2']))
            summary_data = [
                ['Metric', 'Amount'],
                ['Total Income', f"${data['total_income']:,.2f}"],
                ['Total Expenses', f"${data['total_expenses']:,.2f}"],
                ['Net Income', f"${data['net_income']:,.2f}"],
                ['Savings Rate', f"{data['savings_rate']:.1f}%"]
            ]

            summary_table = Table(summary_data, colWidths=[3*inch, 2*inch])
            summary_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1976d2')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 12),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))

            story.append(summary_table)
            story.append(Spacer(1, 0.5 * inch))

            # Spending by category
            if data['spending_by_category']:
                story.append(Paragraph("Spending by Category", styles['Heading2']))

                category_data = [['Category', 'Amount', 'Percentage']]
                for item in data['spending_by_category'][:10]:  # Top 10
                    category_data.append([
                        item['category'],
                        f"${item['amount']:,.2f}",
                        f"{item['percentage']:.1f}%"
                    ])

                category_table = Table(category_data, colWidths=[2.5*inch, 1.5*inch, 1.5*inch])
                category_table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1976d2')),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                    ('ALIGN', (1, 0), (-1, -1), 'RIGHT'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, 0), 10),
                    ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                    ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                    ('GRID', (0, 0), (-1, -1), 1, colors.black)
                ]))

                story.append(category_table)

            # Build PDF
            doc.build(story)

            return {
                "success": True,
                "file_path": output_path,
                "message": "PDF report generated successfully"
            }
        except IOError as e:
            return {
                "success": False,
                "error": f"Failed to generate PDF report: {str(e)}"
            }

    def generate_excel_report(
        self,
        user_id: int,
        report_type: str,
        start_date: datetime,
        end_date: datetime,
        output_path: str
    ) -> Dict[str, Any]:
        """
        Generate Excel report

        Args:
            user_id: User ID
            report_type: Type of report
            start_date: Report start date
            end_date: Report end date
            output_path: Path to save Excel file

        Returns:
            Dict containing generation result
        """
        if not OPENPYXL_AVAILABLE:
            return {
                "success": False,
                "error": "openpyxl not installed. Install with: pip install openpyxl"
            }

        # Validate inputs
        if user_id <= 0:
            return {
                "success": False,
                "error": "Invalid user ID: must be positive"
            }

        if start_date > end_date:
            return {
                "success": False,
                "error": "Start date must be before end date"
            }

        # Validate output directory exists
        output_dir = os.path.dirname(output_path)
        if output_dir and not os.path.exists(output_dir):
            return {
                "success": False,
                "error": f"Output directory does not exist: {output_dir}"
            }

        try:
            # Create workbook
            wb = Workbook()

            # Remove default sheet
            wb.remove(wb.active)

            # Get data
            data = self._get_report_data(user_id, start_date, end_date)

            # Summary sheet
            ws_summary = wb.create_sheet("Summary")
            self._create_summary_sheet(ws_summary, data, start_date, end_date)

            # Transactions sheet
            ws_transactions = wb.create_sheet("Transactions")
            self._create_transactions_sheet(ws_transactions, user_id, start_date, end_date)

            # Category breakdown sheet
            ws_categories = wb.create_sheet("Categories")
            self._create_categories_sheet(ws_categories, data)

            # Save workbook
            wb.save(output_path)

            return {
                "success": True,
                "file_path": output_path,
                "message": "Excel report generated successfully"
            }
        except IOError as e:
            return {
                "success": False,
                "error": f"Failed to generate Excel report: {str(e)}"
            }

    def _get_report_data(
        self,
        user_id: int,
        start_date: datetime,
        end_date: datetime
    ) -> Dict[str, Any]:
        """Get financial data for report"""
        query = text("""
        SELECT
            amount,
            category,
            type,
            date,
            description
        FROM transactions
        WHERE user_id = :user_id
            AND date >= :start_date
            AND date <= :end_date
        """)

        df = load_data_to_dataframe(query, params={
            'user_id': user_id,
            'start_date': start_date.isoformat(),
            'end_date': end_date.isoformat()
        })

        if df.empty:
            return {
                "total_income": 0,
                "total_expenses": 0,
                "net_income": 0,
                "savings_rate": 0,
                "spending_by_category": [],
                "transaction_count": 0
            }

        # Calculate metrics
        income_df = df[df['type'] == 'income']
        expense_df = df[df['type'] == 'expense']

        total_income = float(income_df['amount'].sum())
        total_expenses = float(abs(expense_df['amount'].sum()))
        net_income = total_income - total_expenses
        savings_rate = (net_income / total_income * 100) if total_income > 0 else 0

        # Category breakdown
        category_spending = expense_df.groupby('category')['amount'].sum().abs()
        spending_by_category = []

        for category, amount in category_spending.items():
            percentage = (amount / total_expenses * 100) if total_expenses > 0 else 0
            spending_by_category.append({
                "category": category,
                "amount": float(amount),
                "percentage": float(percentage)
            })

        spending_by_category.sort(key=lambda x: x['amount'], reverse=True)

        return {
            "total_income": total_income,
            "total_expenses": total_expenses,
            "net_income": net_income,
            "savings_rate": savings_rate,
            "spending_by_category": spending_by_category,
            "transaction_count": len(df)
        }

    def _create_summary_sheet(self, ws, data, start_date, end_date):
        """Create summary sheet in Excel"""
        # Title
        ws['A1'] = 'Financial Summary Report'
        ws['A1'].font = Font(size=16, bold=True, color='1976d2')

        # Period
        ws['A2'] = f"Period: {start_date.strftime('%Y-%m-%d')} to {end_date.strftime('%Y-%m-%d')}"

        # Headers
        ws['A4'] = 'Metric'
        ws['B4'] = 'Amount'

        header_fill = PatternFill(start_color='1976d2', end_color='1976d2', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF')

        for cell in ['A4', 'B4']:
            ws[cell].fill = header_fill
            ws[cell].font = header_font

        # Data
        metrics = [
            ('Total Income', data['total_income']),
            ('Total Expenses', data['total_expenses']),
            ('Net Income', data['net_income']),
            ('Savings Rate', f"{data['savings_rate']:.1f}%"),
            ('Transaction Count', data['transaction_count'])
        ]

        for idx, (metric, value) in enumerate(metrics, start=5):
            ws[f'A{idx}'] = metric
            if isinstance(value, (int, float)):
                ws[f'B{idx}'] = value
                ws[f'B{idx}'].number_format = '$#,##0.00'
            else:
                ws[f'B{idx}'] = value

        # Adjust column widths
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 15

    def _create_transactions_sheet(self, ws, user_id, start_date, end_date):
        """Create transactions sheet in Excel"""
        # Get transactions
        query = text("""
        SELECT
            date,
            description,
            category,
            type,
            amount
        FROM transactions
        WHERE user_id = :user_id
            AND date >= :start_date
            AND date <= :end_date
        ORDER BY date DESC
        """)

        df = load_data_to_dataframe(query, params={
            'user_id': user_id,
            'start_date': start_date.isoformat(),
            'end_date': end_date.isoformat()
        })

        # Headers
        headers = ['Date', 'Description', 'Category', 'Type', 'Amount']
        for idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=idx, value=header)
            cell.fill = PatternFill(start_color='1976d2', end_color='1976d2', fill_type='solid')
            cell.font = Font(bold=True, color='FFFFFF')

        # Data
        if not df.empty:
            for row_idx, row in enumerate(df.itertuples(), start=2):
                ws.cell(row=row_idx, column=1, value=row.date)
                ws.cell(row=row_idx, column=2, value=row.description)
                ws.cell(row=row_idx, column=3, value=row.category)
                ws.cell(row=row_idx, column=4, value=row.type)
                ws.cell(row=row_idx, column=5, value=float(row.amount))

        # Format
        for col in range(1, 6):
            ws.column_dimensions[chr(64 + col)].width = 15

    def _create_categories_sheet(self, ws, data):
        """Create categories sheet in Excel"""
        # Headers
        headers = ['Category', 'Amount', 'Percentage']
        for idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=idx, value=header)
            cell.fill = PatternFill(start_color='1976d2', end_color='1976d2', fill_type='solid')
            cell.font = Font(bold=True, color='FFFFFF')

        # Data
        for row_idx, item in enumerate(data['spending_by_category'], start=2):
            ws.cell(row=row_idx, column=1, value=item['category'])
            ws.cell(row=row_idx, column=2, value=item['amount'])
            ws.cell(row=row_idx, column=3, value=item['percentage'])

        # Format
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 15
